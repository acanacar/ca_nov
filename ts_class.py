from constants import *
from ts.add_indicator import *

import requests
import io


# res = requests.get('http://sentdex.com/api/finance/sentiment-signals/sample/')
# if res.status_code is 200:
#     response = res.content
#     # lines = response.splitlines()
#     sentiment_dataframe = pd.read_csv(io.StringIO(response.decode('utf-8')))
#     sentiment_dataframe['date'] = pd.to_datetime(sentiment_dataframe['date'])

class tstable(object):
    algDict = {'logr': logr,
               'knn': knn,
               'svc_poly': svc_poly,
               'svc_linear': svc_linear,
               'svc_sigmoid': svc_sigmoid,
               'gnb': gnb,
               'dtc': dtc,
               'rfc': rfc,
               'lin_reg': linear_reg,
               'svr_rbf_reg': svr_rbf,
               'svr_lin_reg': svr_lin,
               'svr_poly_reg': svr_poly,
               }
    data = pd.read_pickle(hist_pkl)

    def __init__(self, stock):
        self.stock = stock
        data = tstable.data.loc[:, ([self.stock], slice(None))]
        data.columns = data.columns.droplevel(0)
        self.dataframe = data.dropna(axis=1, how='all').dropna(axis=0, how='any')
        self.return_series = self.dataframe['Adj Close'].pct_change().iloc[1:]
        self.indicator_df = self.dataframe.iloc[1:]
        self.classification_df = self.dataframe.iloc[1:]
        self.regression_df = self.dataframe.iloc[1:]

    def add_indicator(self):
        self.indicator_df = add_indicator(data_frame=self.indicator_df)

    def calculate_correlation(self):

        security_tickers = Security_Tickers()
        vix_tickers = Vix_Tickers()
        bond_tickers = Bonds_Tickers()

        return_series = []

        for i in security_tickers:
            data = pd.read_pickle(hist_pkl)
            data = data.loc[:, ([i], ['Adj Close'])]
            return_series.append(data.pct_change().iloc[1:])

        store = pd.HDFStore(hdf5_store)
        for i in vix_tickers:
            data = store['daily/vix']
            data = data.loc[:, ([i], ['Adj Close'])]
            return_series.append(data.pct_change().iloc[1:])
        for i in bond_tickers:
            data = store['daily/usbonds']
            data = data.loc[:, ([i], ['Adj Close'])]
            return_series.append(data.pct_change().iloc[1:])
        if store.is_open:
            store.close()

        return_df = pd.concat(return_series, axis=1)
        return_df.columns = return_df.columns.droplevel(1)
        self.correlation_matrix = return_df.corr()

    def correlation_to_excel(self):
        from openpyxl import load_workbook

        book_path = r'C:\Users\a.acar\PycharmProjects\ca_nov\outputs\book.xlsx'
        book = load_workbook(book_path)

        writer = pd.ExcelWriter(
            path=book_path,
            engine='openpyxl')

        writer.book = book
        self.correlation_matrix.to_excel(writer, sheet_name='Correlation')
        writer.save()
        writer.close()

    def set_classification_label(self):

        def conditions_series(r, yuzde_1=False):
            if yuzde_1 is False:
                if r > 0:
                    return 1
                elif r == 0:
                    return 0
                else:
                    return -1
            else:
                if r >= 0.01:
                    return 1
                elif r <= -0.01:
                    return -1
                else:
                    return 0

        self.classification_df.loc[:, 'label'] = self.return_series.apply(conditions_series, yuzde_1=False)
        self.classification_df.loc[:, 'label_1_pct'] = self.return_series.apply(conditions_series,
                                                                                yuzde_1=True)

    def set_cols_to_classify(self, input_cols):
        self.classifying_input_cols = input_cols
        self.classify_input = self.classification_df.loc[:, self.classifying_input_cols]

    def getScaledData(self, x_train, x_test):
        sc = StandardScaler()
        self.x_train = sc.fit_transform(x_train)
        self.x_test = sc.transform(x_test)

    def set_classification_train_test_data(self):
        X = self.classify_input
        Y = self.classification_df['label']
        x_train, x_test, self.y_train, self.y_test = train_test_split(X, Y,
                                                                      test_size=0.3, random_state=0)
        self.getScaledData(x_train, x_test)

    def set_classifier(self, algo='logr'):
        self.classifier = tstable.algDict[algo]

    def fit_classifier(self):
        self.classifier.fit(self.x_train, self.y_train)

    def make_predictions_classification(self):
        self.y_predictions = self.classifier.predict(self.x_test)

    def set_predictions_to_dataframe_classification(self, algo):
        self.classification_df.loc[-len(self.y_predictions):, algo] = self.y_predictions

    def classify(self, algo, input_cols=['Open', 'High', 'Low', 'Close', 'Volume']):
        self.set_cols_to_classify(input_cols=input_cols)
        self.set_classification_label()
        self.set_classifier(algo=algo)
        self.set_classification_train_test_data()
        self.fit_classifier()
        self.make_predictions_classification()
        self.set_predictions_to_dataframe_classification(algo)

    def external_datas_to_excel(self):
        from openpyxl import load_workbook

        store = pd.HDFStore(hdf5_store)
        '''
        /minute/tickers
        /hour/tickers
        /daily/df_
        /daily/tahvils
        /daily/tickers
        /daily/usbonds
        /daily/vix
        /daily/vix_v2
        '''
        df_vix = pd.concat([store['daily/vix'], store['daily/vix_v2'], store['daily/usbonds']], axis=1)
        store.close()
        book_path = r'C:\Users\a.acar\PycharmProjects\ca_nov\outputs\book.xlsx'
        book = load_workbook(book_path)

        writer = pd.ExcelWriter(
            path=book_path,
            engine='openpyxl')
        writer.book = book

        df_vix.to_excel(writer, sheet_name="VIX and Bonds")
        writer.save()
        writer.close()

    def to_excel(self):
        from openpyxl import load_workbook

        book_path = r'C:\Users\a.acar\PycharmProjects\ca_nov\outputs\book.xlsx'
        book = load_workbook(book_path)

        writer = pd.ExcelWriter(
            path=book_path,
            engine='openpyxl')

        writer.book = book
        df_excel = pd.concat([self.classification_df, self.regression_df, self.indicator_df], axis=1)
        df_excel.to_excel(writer, sheet_name=self.stock)
        writer.save()
        writer.close()

    def set_cols_to_regress(self, input_cols):
        self.regression_input_cols = input_cols
        self.regression_input = self.regression_df.loc[:, self.regression_input_cols]

    def set_dependent_variable(self):
        self.regression_dependent_variable = self.regression_df['Close']

    def set_regression_algo(self, algo):
        self.regression_algo = tstable.algDict[algo]

    def set_regression_train_test_data(self):
        X = self.regression_input
        Y = self.regression_dependent_variable
        self.x_train, self.x_test, self.y_train, self.y_test = train_test_split(X, Y,
                                                                                test_size=0.3, random_state=0)

    def fit_regression(self):
        self.regression_algo.fit(self.x_train, self.y_train)

    def make_predictions_regression(self):
        self.y_predictions = self.regression_algo.predict(self.x_test)

    def set_predictions_to_dataframe_regression(self, algo):
        self.regression_df.loc[-len(self.y_predictions):, algo] = self.y_predictions

    def regression(self, algo, input_cols=['Open', 'High', 'Low', 'Volume']):
        self.set_cols_to_regress(input_cols=input_cols)
        self.set_dependent_variable()
        self.set_regression_algo(algo=algo)
        self.set_regression_train_test_data()
        self.fit_regression()
        self.make_predictions_regression()
        self.set_predictions_to_dataframe_regression(algo)
